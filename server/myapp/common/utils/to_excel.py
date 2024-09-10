from openpyxl import Workbook

def to_excel(d, scale):
    files = d["files"]
    whole_inner_average = d["whole_inner_average"]
    whole_outer_average = d["whole_outer_average"]

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Total inner diameter average'
    ws['C1'] = 'Total outer diameter average'
    ws['E1'] = 'Total ratio average'
    ws['A2'] = whole_inner_average
    ws['C2'] = whole_outer_average
    ws['E2'] = whole_inner_average / whole_outer_average
    ws['A4'] = 'Filename'
    ws['A5'] = 'Inner Average'
    ws['A6'] = 'Outer Average'
    col = ord('C')
    for file in files:
        ws[f'{chr(col)}4'] = file['filename']
        ws[f'{chr(col)}5'] = file['inner_average']
        ws[f'{chr(col)}6'] = file['outer_average']
        ws[f'{chr(col)}8'] = 'Inner'
        ws[f'{chr(col+1)}8'] = 'Outer'
        ws[f'{chr(col+2)}8'] = 'Ratio'
        row = 9
        for result in file["results"]:
            ws[f'{chr(col)}{row}'] = result['inner_diameter']
            ws[f'{chr(col+1)}{row}'] = result['outer_diameter']
            ws[f'{chr(col+2)}{row}'] = result['ratio']
            row += 1

        col += 4
    wb.save('common/tmp/excel/data.xlsx')